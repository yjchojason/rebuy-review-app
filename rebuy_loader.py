from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CAMPAIGN_RE = re.compile(r"^20\d{2}(0[1-9]|1\d|2[0-6])$")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    """Normalize headers for matching, preserving enough meaning."""
    return re.sub(r"\s+", " ", clean_text(value).replace("\n", " ")).strip()


def is_campaign(value: Any) -> bool:
    s = clean_text(value)
    if s.endswith(".0"):
        s = s[:-2]
    return bool(CAMPAIGN_RE.match(s))


def to_campaign(value: Any) -> str:
    s = clean_text(value)
    if s.endswith(".0"):
        s = s[:-2]
    return s


def make_unique(headers: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, h in enumerate(headers, start=1):
        base = clean_text(h) or f"Unnamed_{get_column_letter(i)}"
        if base in seen:
            seen[base] += 1
            out.append(f"{base}__{seen[base]}")
        else:
            seen[base] = 1
            out.append(base)
    return out


def value_to_display(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)


def numeric_or_zero(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def hash_file(path: str | Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def safe_filename(name: str) -> str:
    stem = Path(name).stem
    return re.sub(r"[^A-Za-z0-9_. -]+", "_", stem).strip() or "uploaded_workbook"


@dataclass
class WorkbookData:
    path: str
    workbook_name: str
    workbook_hash: str
    meeting_id: str
    meeting_date: str
    rebuys: pd.DataFrame
    actual_sales: pd.DataFrame
    actual_campaigns: list[str]
    future_campaigns: list[str]
    calendar_map: dict[str, str]
    cm_map: dict[str, list[str]]
    finance_comment_header: str
    finance_comment_column_letter: str
    observed_sheets: dict[str, tuple[int, int]]


FIELD_CANDIDATES = {
    "planner": ["PLANNER"],
    "category": ["CATEGORY"],
    "profile": ["PROFILE #"],
    "fsc": ["FSC#", "FSC"],
    "abc": ["ABC"],
    "coo": ["COO"],
    "global": ["GLOBAL"],
    "full_description": ["FULL DESCRIPTION"],
    "description": ["DESCRIPTION"],
    "std_cost": ["STD COST", "STANDARD COST"],
    "supplier": ["SUPPLIER"],
    "total_dep_demand": ["TTL DEP DMD"],
    "grand_total_demand": ["GRAND TTL DMD"],
    "zan_oh": ["ZAN OH"],
    "mg_oh": ["MG OH"],
    "bk_oh": ["BK OH"],
    "can_oh": ["CAN OH"],
    "pr_oh": ["PR OH"],
    "na_oh": ["NA OH"],
    "na_ait": ["NA AIT"],
    "pos": ["POs", "POS"],
    "bos": ["BOs", "BOS"],
    "cust_orders": ["CUST ORDERS"],
    "total_na_inv": ["TTL NA INV"],
    "safety_stock": ["SS", "SAFETY STOCK"],
    "plc": ["PLC"],
    "coi": ["COI"],
    "dis": ["DIS"],
    "avg_sales_per_camp": ["NA AVG SALES PER CAMP"],
    "rebuy_qty": ["RECO NA REBUY QTY"],
    "frees": ["Frees during 26 campaign period in thousands"],
    "rebuy_dollars": ["REBUY $"],
    "ttl_inv_plus_rebuy": ["TTL NA INV + REBUY"],
    "excess_units": ["EXCESS UNITS"],
    "excess_dollars": ["EXCESS $"],
    "past_26_sales": ["PAST 26 CAMP SALES"],
    "demo_sales": ["DEMO SALES"],
    "stockout_no_ss": ["CAMP OF STOCK OUT W/O SS"],
    "stockout_with_ss": ["CAMP OF STOCK OUT W/ SS"],
    "coverage_camps": ["# OF CAMPS OF COVERAGE"],
    "coverage_campaign": ["CAMP COVERAGE"],
    "fill_by_campaign": ["FB CAMP"],
    "moq": ["MOQ"],
    "moq_coverage": ["MOQ CAMP COVERAGE"],
    "mgt_lt_days": ["MGT LT (DAYS)"],
    "transit_lt_days": ["TRANSIT LT (DAYS)", "TRANSIT LT \n(DAYS)"],
    "grand_lt_days": ["GRAND LT (DAYS)", "GRAND LT\n(DAYS)"],
    "grand_lt_weeks": ["GRAND LT (WKS)", "GRAND LT\n(WKS)"],
    "grand_lt_camps": ["GRAND LT (CAMPS)", "GRAND LT\n(CAMPS)"],
    "planner_comments": ["Planner Commennts", "Planner Comments"],
    "campaign_planning_comments": ["Campaign Planning Comments"],
    "rebuy_meeting_comment": ["Rebuy Meeting Comment"],
    "rebuy_decision": ["Rebuy (Y/N/TBD)", "Rebuy \n(Y/N/TBD)"],
    "initial_finance_comments": ["Initial Finance Comments"],
    "next_12c_actual": ["Finance Team: Next 12C Sales (Based on Actual)", "Finance Team:\nNext 12C Sales\n(Based on Actual)"],
    "next_26c_actual": ["Finance Team: Next 26C Sales (Based on Actual)", "Finance Team:\nNext 26C Sales\n(Based on Actual)"],
    "avg_actual_sales": ["Finance Team: Avg Sales per Campaign (Based on Actual)", "Finance Team:\nAvg Sales \nper Campaign\n(Based on Actual) "],
    "lt_to_deplete": ["Finance Team: LT to Deplete Inv (Campaigns)", "Finance Team:\nLT to Deplete Inv\n(Campaigns)"],
    "updated_fb_lt": ["Updated FB LT (Campaigns)", "Updated FB  LT\n(Campaigns)"],
    "final_finance_reco": ["Final Finance Recommendation", "Final Finance \nRecommendation"],
}


def find_col(df_or_cols, logical_name: str) -> str | None:
    cols = list(df_or_cols.columns) if hasattr(df_or_cols, "columns") else list(df_or_cols)
    norm_to_col = {normalize_header(c).upper(): c for c in cols}
    for candidate in FIELD_CANDIDATES.get(logical_name, []):
        key = normalize_header(candidate).upper()
        if key in norm_to_col:
            return norm_to_col[key]
    return None


def get_value(row: pd.Series, logical_name: str, default: Any = "") -> Any:
    col = find_col(row.index, logical_name)
    if not col:
        return default
    value = row.get(col, default)
    if pd.isna(value) if isinstance(value, float) else False:
        return default
    return value


def load_display_labels(path: str | Path = "config/display_labels.json") -> dict[str, str]:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def parse_rebuys_sheet(wb) -> tuple[pd.DataFrame, list[str], str, str]:
    if "REBUYS" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a REBUYS sheet.")
    ws = wb["REBUYS"]
    row_iter = ws.iter_rows(min_row=1, max_row=1000, max_col=ws.max_column, values_only=True)
    raw_headers = list(next(row_iter))
    headers = make_unique(raw_headers)
    future_campaigns = [to_campaign(h) for h in raw_headers if is_campaign(h)]

    data = []
    empty_run = 0
    for excel_row, values in enumerate(row_iter, start=2):
        fsc = values[3] if len(values) >= 4 else None
        if fsc in (None, ""):
            empty_run += 1
            if data and empty_run >= 25:
                break
            continue
        empty_run = 0
        row = dict(zip(headers, values))
        row["__row_number"] = excel_row
        data.append(row)

    df = pd.DataFrame(data)
    if df.empty:
        raise ValueError("REBUYS sheet was found, but no data rows with FSC# were detected.")

    # Find the Initial Finance Comments column dynamically. In this file it is CM,
    # but future versions may shift.
    finance_header = None
    finance_letter = "CM"
    for idx, h in enumerate(raw_headers, start=1):
        if normalize_header(h).upper() == "INITIAL FINANCE COMMENTS":
            finance_header = headers[idx - 1]
            finance_letter = get_column_letter(idx)
            break
    if not finance_header:
        raise ValueError("Could not find the 'Initial Finance Comments' column in REBUYS.")

    return df, future_campaigns, finance_header, finance_letter


def parse_calendar(wb) -> tuple[dict[str, str], str]:
    if "CALENDAR" not in wb.sheetnames:
        return {}, ""
    ws = wb["CALENDAR"]
    calendar_map: dict[str, str] = {}
    meeting_date = ""
    # The file stores today's/meeting date in G2.
    g2 = ws["G2"].value
    if isinstance(g2, datetime):
        meeting_date = g2.strftime("%Y-%m-%d")
    elif g2:
        meeting_date = clean_text(g2)

    for row in ws.iter_rows(min_row=3, max_col=3, values_only=True):
        campaign = to_campaign(row[1])
        date_value = row[2]
        if is_campaign(campaign):
            if isinstance(date_value, datetime):
                calendar_map[campaign] = date_value.strftime("%Y-%m-%d")
            else:
                calendar_map[campaign] = clean_text(date_value)
    return calendar_map, meeting_date


def parse_actual_sales(wb) -> tuple[pd.DataFrame, list[str]]:
    if "ACTUAL SALES" not in wb.sheetnames:
        return pd.DataFrame(), []
    ws = wb["ACTUAL SALES"]
    campaign_values = [cell for cell in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    campaigns_by_idx: dict[int, str] = {}
    for idx, value in enumerate(campaign_values, start=1):
        if is_campaign(value):
            campaigns_by_idx[idx] = to_campaign(value)

    if not campaigns_by_idx:
        return pd.DataFrame(), []

    max_needed_col = max(max(campaigns_by_idx), 3)
    rows = []
    # Data begins in row 4. Stop at the physical used rows for ACTUAL SALES.
    for values in ws.iter_rows(min_row=4, max_col=max_needed_col, values_only=True):
        fsc = clean_text(values[0] if len(values) >= 1 else "")
        if not fsc:
            continue
        out = {"FSC": fsc, "Description": values[2] if len(values) >= 3 else ""}
        for idx, camp in campaigns_by_idx.items():
            out[camp] = numeric_or_zero(values[idx - 1] if idx - 1 < len(values) else 0)
        rows.append(out)
    if not rows:
        return pd.DataFrame(), list(campaigns_by_idx.values())
    df = pd.DataFrame(rows)
    campaign_cols = list(campaigns_by_idx.values())
    # Some ACTUAL SALES tabs can contain multiple rows per FSC; aggregate just in case.
    agg = {c: "sum" for c in campaign_cols}
    agg["Description"] = "first"
    df = df.groupby("FSC", as_index=False).agg(agg)
    return df, campaign_cols


def parse_cm_sku(wb) -> dict[str, list[str]]:
    if "CM SKU" not in wb.sheetnames:
        return {}
    ws = wb["CM SKU"]
    links: dict[str, set[str]] = {}
    for old_fsc, _, new_fsc, *_ in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        old_s = clean_text(old_fsc)
        new_s = clean_text(new_fsc)
        if not old_s or not new_s or old_s == new_s:
            continue
        links.setdefault(old_s, set()).add(new_s)
        links.setdefault(new_s, set()).add(old_s)
    return {k: sorted(v) for k, v in links.items()}


def load_workbook_data(path: str | Path, workbook_name: str | None = None) -> WorkbookData:
    path = str(path)
    workbook_name = workbook_name or Path(path).name
    workbook_hash = hash_file(path)
    wb = load_workbook(path, data_only=True, read_only=True)

    observed_sheets = {ws.title: (ws.max_row, ws.max_column) for ws in wb.worksheets}
    rebuys, future_campaigns, finance_header, finance_col_letter = parse_rebuys_sheet(wb)
    actual_sales, actual_campaigns = parse_actual_sales(wb)
    calendar_map, meeting_date = parse_calendar(wb)
    cm_map = parse_cm_sku(wb)

    if not meeting_date:
        # Fallback to date in the filename or today's date.
        match = re.search(r"(\d{1,2})[-_. ](\d{1,2})[-_. ](\d{2,4})", workbook_name)
        if match:
            mm, dd, yy = match.groups()
            yy = ("20" + yy) if len(yy) == 2 else yy
            meeting_date = f"{yy}-{int(mm):02d}-{int(dd):02d}"
        else:
            meeting_date = datetime.now().strftime("%Y-%m-%d")

    # Meeting id is stable enough for a bi-weekly file, but includes hash prefix to avoid accidental mixing.
    meeting_id = f"{meeting_date}__{safe_filename(workbook_name)}__{workbook_hash[:10]}"
    return WorkbookData(
        path=path,
        workbook_name=workbook_name,
        workbook_hash=workbook_hash,
        meeting_id=meeting_id,
        meeting_date=meeting_date,
        rebuys=rebuys,
        actual_sales=actual_sales,
        actual_campaigns=actual_campaigns,
        future_campaigns=future_campaigns,
        calendar_map=calendar_map,
        cm_map=cm_map,
        finance_comment_header=finance_header,
        finance_comment_column_letter=finance_col_letter,
        observed_sheets=observed_sheets,
    )


def actual_sales_long(data: WorkbookData, fsc: str, include_related: bool = True) -> pd.DataFrame:
    if data.actual_sales.empty or not data.actual_campaigns:
        return pd.DataFrame(columns=["Campaign", "US Start Date", "FSC", "Sales Qty"])
    fscs = [clean_text(fsc)]
    if include_related:
        fscs.extend(data.cm_map.get(clean_text(fsc), []))
    fscs = [x for x in dict.fromkeys(fscs) if x]
    subset = data.actual_sales[data.actual_sales["FSC"].isin(fscs)]
    if subset.empty:
        return pd.DataFrame(columns=["Campaign", "US Start Date", "FSC", "Sales Qty"])
    long = subset.melt(id_vars=["FSC"], value_vars=data.actual_campaigns, var_name="Campaign", value_name="Sales Qty")
    long["US Start Date"] = long["Campaign"].map(data.calendar_map).fillna("")
    long["Sales Qty"] = pd.to_numeric(long["Sales Qty"], errors="coerce").fillna(0)
    return long[["Campaign", "US Start Date", "FSC", "Sales Qty"]]


def future_demand_long(data: WorkbookData, row: pd.Series) -> pd.DataFrame:
    records = []
    for camp in data.future_campaigns:
        if camp in row.index:
            value = row.get(camp, 0)
        else:
            value = 0
        records.append(
            {
                "Campaign": camp,
                "US Start Date": data.calendar_map.get(camp, ""),
                "Future Demand": numeric_or_zero(value),
            }
        )
    return pd.DataFrame(records)


def export_workbook_with_comments(
    source_path: str | Path,
    comments: pd.DataFrame,
    output_path: str | Path,
    finance_header_text: str = "Initial Finance Comments",
) -> str:
    """Write locally stored Finance comments back into the original workbook.

    Uses the original row number and validates FSC in column D before writing.
    """
    wb = load_workbook(source_path, data_only=False)
    if "REBUYS" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a REBUYS sheet.")
    ws = wb["REBUYS"]

    finance_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if normalize_header(cell.value).upper() == normalize_header(finance_header_text).upper():
            finance_col = idx
            break
    if finance_col is None:
        # Conservative fallback to CM for the current workbook design.
        finance_col = 91

    for _, rec in comments.iterrows():
        comment = clean_text(rec.get("Finance Comment", ""))
        if not comment:
            continue
        row_num = int(rec["Original Row Number"])
        fsc_from_sheet = clean_text(ws.cell(row=row_num, column=4).value)
        fsc_from_comment = clean_text(rec.get("FSC", ""))
        # Write only when row still appears to match. This prevents accidental writes
        # if someone exports against the wrong workbook.
        if fsc_from_sheet and fsc_from_comment and fsc_from_sheet != fsc_from_comment:
            continue
        ws.cell(row=row_num, column=finance_col).value = comment

    output_path = str(output_path)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
